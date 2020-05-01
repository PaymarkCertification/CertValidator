import pandas as pd
import numpy as np
import sys
from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import get_column_letter
import csv

def paymentsNZ():
    driver = webdriver.Chrome()
    print("Init Chrome browser")
    driver.get('https://www.paymentsnz.co.nz/resources/industry-registers/device-register/')
    try:
        WebDriverWait(driver, 5).\
             until(EC.presence_of_element_located((By.CSS_SELECTOR, "body")))
        driver.find_element_by_css_selector('body')
        print("Element loaded")

        table = driver.find_element_by_class_name('viewport')
        head = table.find_element_by_tag_name('thead')
        body = table.find_element_by_tag_name('tbody')
        print('Elements located')
    except NoSuchElementException as e:
        print('Unable to locate element', e)
        exit()

    fullList = open('PNZ.txt', 'w')
    i = 0
    rows = body.find_elements(By.TAG_NAME, "tr") # get all of the rows in the table
    print("Retrieving all table values")
    for row in rows:
        x = 1
        i = i+x
        # Get the columns (all the column 2)

        manu = row.find_elements(By.TAG_NAME,    "td")[0]#note: index start from 0, 1 is col 2
        pci  = row.find_elements(By.TAG_NAME,    "td")[1]
        appr = row.find_elements(By.TAG_NAME,    "td")[2]
        date = row.find_elements(By.TAG_NAME,    "td")[3]
        pnzDate = row.find_elements(By.TAG_NAME, "td")[4]

        print(manu.text + '\\' + pci.text + '\\' + appr.text + '\\' + date.text + '\\' + pnzDate.text, file=fullList)
    print("Rows retrieved: ", i)
    print("\nExport to text")
    driver.quit()
    print("Closing browser")

# paymentsNZ()

excel = "datafile.xlsx"
wb = Workbook()
ws = wb.active
f = open('PNZ.txt')
csv.register_dialect('slash', delimiter='\\')
reader = csv.reader(f, dialect='slash')
wb = Workbook()
ws = wb.worksheets[0]
for row_index, row in enumerate(reader):
    for column_index, cell in enumerate(row):
        column_letter = get_column_letter((column_index + 1))
        ws['%s%s' % (column_letter, (row_index + 1))].value = cell

ws.insert_rows(1)
a = ws['A1'] = "MANUFACTURER & MODEL"
b = ws['B1'] = "PCI APPROVAL NUMBER"
c = ws['C1'] = "APPROVAL VERSION/CLASS"
d = ws['D1'] = "PAYMENTS NZ DATE OF NO NEW CONNECTION"
e = ws['E1'] = "PAYMENTS NZ SUNSET DATE"
wb.save(excel)


df = pd.read_excel(excel)
print("loading into pandas")
ndf = pd.DataFrame(df, columns=['MANUFACTURER & MODEL', 'PCI APPROVAL NUMBER', 'APPROVAL VERSION/CLASS',
                   'PAYMENTS NZ DATE OF NO NEW CONNECTION', 'PAYMENTS NZ SUNSET DATE'])
ndf.dropna(subset=['MANUFACTURER & MODEL'], inplace=True)
ndf[ndf.'MANUFACTURER & MODEL' != 'Show me more']
print(ndf)
writer = pd.ExcelWriter(excel, engine='xlsxwriter')
df.to_excel(writer, sheet_name='sheet1', index=False)
writer.save()
print("saving file")
